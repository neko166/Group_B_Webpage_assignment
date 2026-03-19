import io
import json
import os
import re
import smtplib
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel as PydanticBase
from sqlalchemy.orm import Session, joinedload

from database import get_db
import models
from routers.roadmap import _call_ollama

router = APIRouter(prefix="/api/roadmap", tags=["summary"])


# ══════════════════════════════════════════════
# スキルサマリー強み自動生成
# ══════════════════════════════════════════════
@router.get("/skill-summary/{user_id}")
async def get_skill_summary(user_id: int, db: Session = Depends(get_db)):
    """職務経歴・スキルを元にLLMで強み・特徴を自動生成して返す"""
    user = db.query(models.User).filter(models.User.user_id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    careers = (
        db.query(models.CareerHistory)
        .filter(models.CareerHistory.user_id == user_id)
        .order_by(models.CareerHistory.period_start.desc())
        .all()
    )
    career_text = "\n".join([
        f"- {c.project_name}（{c.role}）: {c.tech_stack} / {c.description}"
        for c in careers
    ]) or "職務経歴未登録"

    user_skills = (
        db.query(models.UserSkill)
        .options(joinedload(models.UserSkill.skill))
        .filter(models.UserSkill.user_id == user_id)
        .all()
    )
    skills_text = "\n".join([
        f"- {us.skill.skill_name}（{us.skill.category}）: レベル {us.skill_level}/5"
        for us in user_skills
    ]) or "スキル未登録"

    prompt = f"""あなたはエンジニアのキャリアコンサルタントです。
以下のエンジニアの職務経歴とスキルを分析し、このエンジニアの強み・特徴を3〜5箇条で簡潔に列挙してください。

【職務経歴】
{career_text}

【保有スキル】
{skills_text}

【出力形式】
JSON配列のみで回答してください。説明文や```は不要です。
["強み1", "強み2", "強み3"]

【ルール】
- 必ず日本語で記述すること
- 各項目は30字以内の簡潔な文にすること
- 職務経歴から読み取れる具体的な実績・経験を反映すること"""

    try:
        raw = await _call_ollama(prompt)
        match = re.search(r'\[[\s\S]*?\]', raw)
        if not match:
            raise ValueError("LLMがJSON配列で返しませんでした")
        strengths = json.loads(match.group())
        if not isinstance(strengths, list):
            raise ValueError("配列ではありません")
    except Exception as e:
        print(f"[skill-summary] LLM error: {e}")
        # OllamaなしでもスキルDBと職務経歴から強みを生成
        strengths = []
        top_skills = [us for us in user_skills if us.skill_level >= 4]
        if top_skills:
            names = "・".join(us.skill.skill_name for us in top_skills[:3])
            strengths.append(f"{names}を活用した即戦力開発")
        cats = list(dict.fromkeys(us.skill.category for us in user_skills))
        if cats:
            strengths.append(f"{' / '.join(cats[:3])}領域の幅広い技術経験")
        if careers:
            roles = list(dict.fromkeys(c.role for c in careers))
            strengths.append(f"{roles[0]}として{len(careers)}プロジェクトの実務経験")
            recent = careers[0]
            techs = [t.strip() for t in (recent.tech_stack or '').split(',') if t.strip()]
            if techs:
                strengths.append(f"直近プロジェクトで{' / '.join(techs[:3])}を使用")
        if not strengths:
            strengths = ["スキルと職務経歴を登録すると詳細な強みが表示されます"]

    return {"strengths": strengths}


# ══════════════════════════════════════════════
# メール送信
# ══════════════════════════════════════════════
class SendEmailRequest(PydanticBase):
    user_id:    int
    to_email:   str
    tab:        str = "proposal"  # "proposal" or "assign"


@router.post("/send-email")
def send_summary_email(payload: SendEmailRequest, db: Session = Depends(get_db)):
    """営業サマリーをメールで送信する"""
    user = db.query(models.User).filter(models.User.user_id == payload.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # スキル取得
    user_skills = (
        db.query(models.UserSkill)
        .options(joinedload(models.UserSkill.skill))
        .filter(models.UserSkill.user_id == payload.user_id)
        .all()
    )

    # 職務経歴取得
    careers = (
        db.query(models.CareerHistory)
        .filter(models.CareerHistory.user_id == payload.user_id)
        .order_by(models.CareerHistory.period_start.desc())
        .all()
    )

    # 最新ロードマップ取得してマッチング案件を検索
    roadmap = (
        db.query(models.Roadmap)
        .filter(models.Roadmap.user_id == payload.user_id)
        .order_by(models.Roadmap.created_at.desc())
        .first()
    )
    match_projects = []
    if roadmap:
        content = json.loads(roadmap.content_json)
        steps = content.get("steps", [])
        current_step = next((s for s in steps if s["status"] == "current"), None) \
                    or next((s for s in steps if s["status"] == "completed"), None) \
                    or (steps[0] if steps else None)
        if current_step:
            step_skills = [s.lower() for s in current_step.get("skills_to_acquire", [])]
            user_skill_names = [us.skill.skill_name.lower() for us in user_skills]
            skill_list = step_skills + user_skill_names
            all_projects = db.query(models.Project).all()
            scored = []
            for p in all_projects:
                req = [s.strip().lower() for s in (p.required_skills or "").split(",") if s.strip()]
                if req:
                    matched = sum(1 for ps in req if any(ps in ss or ss in ps for ss in skill_list))
                    score = int(matched / len(req) * 100)
                else:
                    score = 0
                scored.append((p, score))
            scored.sort(key=lambda x: x[1], reverse=True)
            match_projects = [(p, s) for p, s in scored[:2]]

    # ── Excelファイル生成 ──
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    engineer_name = user.name or "エンジニア"
    wb = openpyxl.Workbook()

    # ヘッダースタイル
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="1A1A2E")
    sub_fill      = PatternFill("solid", fgColor="3B5BDB")
    center        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left          = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    thin          = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def set_header(ws, row, col, text, width=None):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = header_font
        cell.fill      = sub_fill
        cell.alignment = center
        cell.border    = thin
        if width:
            ws.column_dimensions[cell.column_letter].width = width
        return cell

    def set_cell(ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = left
        cell.border    = thin
        return cell

    # ── シート1: プロフィール ──
    ws1 = wb.active
    ws1.title = "エンジニアプロフィール"
    ws1.row_dimensions[1].height = 30
    title_cell = ws1.cell(row=1, column=1, value=f"営業向けサマリー：{engineer_name}")
    title_cell.font      = Font(bold=True, size=14, color="1A1A2E")
    title_cell.alignment = left
    ws1.merge_cells("A1:B1")

    labels = ["氏名", "年齢", "経験年数", "現在のロール", "目標ロール", "稼働可能時期", "勤務地"]
    values = [
        user.name or "—",
        f"{user.age}歳" if user.age else "—",
        f"{user.experience_years}年" if user.experience_years else "—",
        user.current_role or "—",
        user.target_role or "—",
        user.available_from or "—",
        user.work_locations or "—",
    ]
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 36
    for i, (label, value) in enumerate(zip(labels, values), start=3):
        lc = ws1.cell(row=i, column=1, value=label)
        lc.font      = Font(bold=True)
        lc.fill      = PatternFill("solid", fgColor="E8EAF6")
        lc.alignment = left
        lc.border    = thin
        vc = ws1.cell(row=i, column=2, value=value)
        vc.alignment = left
        vc.border    = thin

    # ── シート2: 保有スキル ──
    ws2 = wb.create_sheet("保有スキル")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    set_header(ws2, 1, 1, "スキル名")
    set_header(ws2, 1, 2, "カテゴリ")
    set_header(ws2, 1, 3, "レベル（/5）")
    set_header(ws2, 1, 4, "評価")
    for i, us in enumerate(user_skills, start=2):
        set_cell(ws2, i, 1, us.skill.skill_name)
        set_cell(ws2, i, 2, us.skill.category or "")
        set_cell(ws2, i, 3, us.skill_level)
        set_cell(ws2, i, 4, "★" * us.skill_level + "☆" * (5 - us.skill_level))

    # ── シート3: プロジェクト実績 ──
    ws3 = wb.create_sheet("プロジェクト実績")
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 24
    ws3.column_dimensions["D"].width = 30
    ws3.column_dimensions["E"].width = 40
    set_header(ws3, 1, 1, "期間")
    set_header(ws3, 1, 2, "プロジェクト名")
    set_header(ws3, 1, 3, "役割")
    set_header(ws3, 1, 4, "技術スタック")
    set_header(ws3, 1, 5, "概要")
    for i, c in enumerate(careers, start=2):
        start_str = c.period_start.strftime("%Y/%m") if c.period_start else ""
        end_str   = c.period_end.strftime("%Y/%m") if c.period_end else "現在"
        period    = f"{start_str} - {end_str}" if start_str else end_str
        set_cell(ws3, i, 1, period)
        set_cell(ws3, i, 2, c.project_name)
        set_cell(ws3, i, 3, c.role or "")
        set_cell(ws3, i, 4, c.tech_stack or "")
        set_cell(ws3, i, 5, c.description or "")
        ws3.row_dimensions[i].height = 36

    # ── シート4: マッチング案件提案 ──
    ws4 = wb.create_sheet("マッチング案件提案")
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 24
    ws4.column_dimensions["C"].width = 12
    ws4.column_dimensions["D"].width = 36
    ws4.column_dimensions["E"].width = 24
    ws4.column_dimensions["F"].width = 16
    set_header(ws4, 1, 1, "案件名")
    set_header(ws4, 1, 2, "企業名")
    set_header(ws4, 1, 3, "マッチ度")
    set_header(ws4, 1, 4, "概要")
    set_header(ws4, 1, 5, "必要スキル")
    set_header(ws4, 1, 6, "雇用形態")
    if match_projects:
        for i, (p, score) in enumerate(match_projects, start=2):
            set_cell(ws4, i, 1, p.project_name)
            set_cell(ws4, i, 2, p.company)
            c3 = set_cell(ws4, i, 3, f"{score}%")
            c3.alignment = center
            c3.font = Font(bold=True, color="3B5BDB")
            set_cell(ws4, i, 4, p.project_overview or "")
            set_cell(ws4, i, 5, p.required_skills or "")
            set_cell(ws4, i, 6, p.employ_type or "")
    else:
        ws4.cell(row=2, column=1, value="ロードマップを生成すると案件提案が表示されます").font = Font(color="AAAAAA")

    # Excelをバイト列に変換
    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_buf.seek(0)
    xlsx_bytes = xlsx_buf.read()

    # ── SMTP送信 ──
    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASSWORD", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user)

    if not smtp_host or not smtp_user or not smtp_pass:
        raise HTTPException(
            status_code=503,
            detail="SMTP設定が未設定です。.env に SMTP_HOST / SMTP_USER / SMTP_PASSWORD を設定してください。"
        )

    subject = f"【営業向けサマリー】{engineer_name} エンジニアのご提案"
    body_text = (
        f"{engineer_name} エンジニアの営業向けサマリーをExcelで添付いたします。\n\n"
        f"■ 現在のロール: {user.current_role or '—'}\n"
        f"■ 目標ロール:   {user.target_role or '—'}\n\n"
        "添付のExcelファイルをご確認ください。\n\n"
        "---\nこのメールはシステムから自動送信されました。"
    )

    try:
        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"]    = smtp_from
        msg["To"]      = payload.to_email
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        filename = f"summary_{engineer_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        xlsx_part = MIMEApplication(xlsx_bytes, _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        xlsx_part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(xlsx_part)

        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_from, payload.to_email, msg.as_string())
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=503, detail="SMTP認証エラー。メールアドレスとパスワードを確認してください。")
    except (smtplib.SMTPConnectError, smtplib.SMTPServerDisconnected, smtplib.SMTPDataError, OSError) as e:
        raise HTTPException(
            status_code=503,
            detail=(
                "SMTPネットワークエラー。ホスト、ポート、インターネット接続、TLS/SSL設定、ファイアウォールを確認してください。"
                f" 詳細: {e}"
            ),
        )
    except smtplib.SMTPException as e:
        raise HTTPException(status_code=503, detail=f"メール送信エラー: {e}")
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"送信失敗: {e}")

    return {"message": f"{payload.to_email} にメールを送信しました"}
